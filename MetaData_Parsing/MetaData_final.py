import xml.etree.ElementTree as elemTree
import openpyxl
import numpy as np
from openpyxl.styles import PatternFill, Color
from collections import OrderedDict
import copy

xmlFilePath, serverFilePath = [], []
only_Server, only_Xml = [], []  # 엑셀에 넣을 배열들
duplication = OrderedDict()
only_s = dict()
only_x = dict()
flag = 0
#파일마다 바꿔주어야할것 : 각각의 파일 Path(xmlPath, serverPath변수,
xmlPath, serverPath = 0,0
xmlRealPath, serverReaPath = [], []
list_for_duplication=[]
duplication_dictionary = OrderedDict()
xml_dictionary_for_index = OrderedDict()
try:
    xmlPath = "D:/minwoo/Meta/MOTIE_20190103.xml"
    targetXML = open(xmlPath, 'r', encoding='UTF8')
#    targetXML = open('MOTIE_20190103.xml','r',encoding='UTF8')
    # tree = elemTree.parse("D:/minwoo/Meta/MOTIE_20190103.xml") #파일 불러오고
    tree = elemTree.parse(targetXML)
    root = tree.getroot()
    # root 가져온다.

    for ele in root.findall("./project_set/first_data/register_file/file"):
        #(trash, item) = ele.get('path').split('igda/')
        xmlRealPath.append(ele.get('path')) #original 경로 저장
        (trash,item) = ele.get('path').rsplit('/',1)
        xmlFilePath.append(item)
        only_x[item] = ele.get('path')
        #onlyx에 경로이름값-실제 path값 저장
        # xmlFilePath.append(ele.get('path'))
        # name만 가져와서 xmlFilePath배열에 저장
        # print(xmlFilePath) #디버깅용


except IOError as err:
    print("Xml File Error : " + str(err))
    # 여기까지 왔으면 xmlFilePath에 xml 파일리스트 저장완료


try:  # Server 파일목록 텍스트파일 읽어와서 파일명 파싱해서 저장
    serverPath = "D:/minwoo/Meta/motie.txt"
    targetServer = open(serverPath)
    lines = targetServer.readlines()

    for line in lines:
        serverReaPath.append(line) #서버 original 경로 저장
        (trash, value) = line.rsplit('/',1)  # 이부분은 파일마다 수정해줘야한다.
        serverFilePath.append(value.rstrip('\n'))  # 개행문자 제거하고 추가
        only_s[value.rstrip('\n')] = line

except IOError as err:
    print("Server File Error: " + str(err))
    # 여기까지 왔으면 serverFilePath에 서버 파일리스트 저장완료

total_len = len(serverFilePath)+len(xmlFilePath)
# 우선 서버정보를 모르니, 서버의 파일경로가 serverFilePath에 저장되어있다고 가정
xmlFilePath.sort()
serverFilePath.sort()
serverReaPath.sort()
xmlRealPath.sort()
temp = sorted(xmlRealPath, key=lambda path: path.rsplit('/',1)[1] )
xmlRealPath = temp
#xmlRealPath.sort()
# 정렬
(trash, temp) = serverPath.split('Meta/')
(output_fileName, temp) = temp.split('.txt')
print(output_fileName)  # outputFile 이름 저장


if np.array_equal(xmlFilePath, serverFilePath):
    flag = 1
    # 두 xml과 서버 값이 동일한경우 flag값 변경(이름지을때 사용)

#only_Server = [x for x in serverFilePath if x not in xmlFilePath]
# 서버에있고, xml에 없는 목록 저장

for x in serverReaPath:
    (trash,temp) = x.rsplit('/',1)
    if temp.rstrip('\n') not in xmlFilePath:
        only_Server.append(x)
    #개선한버전

#only_Xml = [x for x in xmlFilePath if x not in serverFilePath]
# xml에 있고, 서버에 없는 목록 저장

for x in xmlRealPath:
    (trash,temp) = x.rsplit('/',1)
    if temp not in serverFilePath:
        only_Xml.append(x)
    #개선한버전

for list in xmlFilePath:
    if list not in duplication.keys():
        duplication[list] = 1
    else:
        duplication[list] += 1
#duplication dictionary에 각 리스트별로 몇개가 중복되는지 입력한다.

for key,value in duplication.items():
    if int(value)>1:
        list_for_duplication.append(key)
        #중복되는 목록만 list_for_duplication에 삽입


for item in list_for_duplication:
    for path in xmlRealPath:
        (trash,temp) = path.rsplit('/',1) #temp에 이름파싱해서임시저장
        if item==temp:
            key = xmlRealPath.index(path)
            duplication_dictionary[key]=path
            #중복이 되는경우를 index값을 key에 주소값을 value에 저장

# 결과값 엑셀파일에 쓰기
wb = openpyxl.Workbook()
sheet = wb.active

sheet['A1'] = '서버 데이터'
sheet['B1'] = 'XML 데이터'
sheet['C1'] = '서버에만 있는 데이터'
sheet['D1'] = 'XML에만 있는 데이터'
sheet['E1'] = 'XML 파일명 중복 여부'


''''
index_server = 0
index_xml = 0
i = 0
print(len(serverReaPath))
print(len(serverFilePath))
while total_len > i :
        if index_server >= len(serverFilePath):
            if index_xml >= len(xmlFilePath):
                break
            else:
                sheet['B' + str(index_server+2)] = xmlFilePath[index_xml]
                index_xml += 1
                index_server += 1
                i += 1
        elif serverFilePath[index_server]==xmlFilePath[index_xml]:
            sheet['A' + str(index_server+2)] = serverFilePath[index_server]
            sheet['B' + str(index_server+2)] = xmlFilePath[index_xml]
            index_server += 1
            index_xml += 1
            i += 1
        else :
            sheet['A' + str(index_server+2)] = serverFilePath[index_server]
            sheet['C' + str(index_server+2)] = only_s[serverFilePath[index_server]]
            index_server += 1
            i += 1
'''
index_server = 0
index_xml = 0
i = 0
temp_xml = copy.deepcopy(xmlFilePath)
while total_len > i :
        if index_server >= len(serverFilePath):
            #if index_xml >= len(xmlFilePath):
            break
            #else:
            #    po = xmlFilePath.pop(0)
            #    sheet['A' + str(index_server)].fill= PatternFill(patternType='solid', fgColor=Color('FAC000'))
            #    sheet['A' + str(index_server)] = 'NONE'
            #    sheet['B' + str(index_server+2)] = po
            #    sheet['D' + str(index_server+2)] = only_x[po]
            #    #xml에만있는거 경로 출력하기
            #    index_xml += 1
            #    index_server += 1
            #    i += 1
        elif serverFilePath[index_server] in xmlFilePath:
            sheet['A' + str(index_server+2)] = serverFilePath[index_server]
            sheet['B' + str(index_server+2)] = serverFilePath[index_server]
            if int(duplication[serverFilePath[index_server]]) > 1:
                sheet['E' + str(index_server+2)].fill= PatternFill(patternType='solid', fgColor=Color('FAC000'))
                sheet['E' + str(index_server+2)] = '중복'
            else:
                sheet['E' + str(index_server+2)] = '유일'
            #xmlFilePath.remove(serverFilePath[index_server])
            index_server += 1
            index_xml += 1
            i += 1
        else :
            sheet['A' + str(index_server+2)] = serverFilePath[index_server]
            sheet['B' + str(index_server+2)].fill = PatternFill(patternType='solid', fgColor=Color('FAC000'))
            sheet['B' + str(index_server+2)] = 'NONE'
            sheet['C' + str(index_server+2)] = only_s[serverFilePath[index_server]]
            index_server += 1
            i += 1


for item in only_Xml:
    sheet['A'+str(index_server+2)] = 'None'
    sheet['A'+str(index_server+2)].fill = PatternFill(patternType='solid', fgColor=Color('FAC000'))
    sheet['B'+str(index_server+2)] = item.rsplit('/',1)[1]
    sheet['D'+str(index_server+2)] = item
    k=item.rsplit('/',1)[1]
    if int(duplication[k]) > 1:
        sheet['E' + str(index_server+2)].fill= PatternFill(patternType='solid', fgColor=Color('FAC000'))
        sheet['E' + str(index_server+2)] = '중복'
    else:
        sheet['E' + str(index_server+2)] = '유일'
    index_server +=1
    # 네번째 칼럼에 only_xml 입력

#k = 'tmp'
#while len(temp_xml) != 0:
    #sheet['A' + str(index_server+2)] = 'None'
#    t = temp_xml.pop(0)
#    sheet['B' + str(index_server+2)] = t
#    sheet['D' + str(index_server+2)] = t

#for key,value in duplication_dictionary.items():
#    sheet['E'+ str(int(key)+2)] = value

'''
index = 2
for item in serverFilePath:
    sheet['A' + str(index)] = item
    index += 1
    # 첫번째 칼럼에 serverFilepath 입력

index = 2
for item in xmlFilePath:
    sheet['B' + str(index)] = item
    index += 1
    # 두번째 칼럼에 xmlFilepath 입력

index = 2
for item in only_Server:
    sheet['C' + str(index)] = item
    index += 1
    # 세번째 칼럼에 only_Server 입력

index = 2
for item in only_Xml:
    sheet['D' + str(index)] = item
    index += 1
    # 네번째 칼럼에 only_xml 입력

index = 2
for item in duplication.keys():
    sheet['E' + str(index)] = item
    index += 1
    # 다섯번째 칼럼에 중복 제거된 xml 리스트 입력

index = 2
for item in duplication.values():
    sheet['F' + str(index)] = item
    index += 1
    #여섯번째 칼럼에 중복 제거된 xml 리스트 각각의 개수 입력
'''

if(flag == 1):
    wb.save(output_fileName + '_일치.xlsx')
else:
    wb.save(output_fileName + '_불일치.xlsx')
